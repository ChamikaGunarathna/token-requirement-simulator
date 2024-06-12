from openpyxl import Workbook
from openpyxl import load_workbook
import os

class ExcelWriter:
    def __init__(self, filename):
        self.filename = filename
        self.fields = [
            "Matches",
            "min_num_players",
            "max_num_players",
            "max_score_points",
            "low_scoring_player_mean_percentage",
            "low_scoring_player_variance_percentage",
            "high_scoring_player_mean_percentage",
            "high_scoring_player_variance_percentage",
            "xp_to_sprt_ratio",
            "total_players",
            "total_score",
            "total_tokens",
            "average_score_per_player",
            "average_tokens_per_player",
            "average_score_per_match",
            "average_tokens_per_match"
            ]

        # Check if file exists, if not, create a new file
        if not os.path.exists(self.filename):
            self.workbook = Workbook()
            self.sheet = self.workbook.active
            self.sheet.append(self.fields)
        else:
            self.workbook = load_workbook(self.filename)
            self.sheet = self.workbook.active

    def save_excel(self, results):
        self.sheet.append(results)
        self.workbook.save(self.filename)